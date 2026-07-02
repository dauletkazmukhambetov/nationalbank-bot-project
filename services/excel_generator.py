from __future__ import annotations

import re
import tempfile
from io import BytesIO
from pathlib import Path
from typing import Any

import matplotlib.pyplot as plt
import pandas as pd
from openpyxl.drawing.image import Image
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from services.report_filters import filter_dataframe


def safe_filename(text: str) -> str:
    text = re.sub(r"[^0-9A-Za-zА-Яа-я_. -]+", "_", text)
    text = re.sub(r"\s+", "_", text).strip("_ .")
    return text[:120] or "report"


def find_records(payload: Any) -> Any:
    if isinstance(payload, list):
        return payload

    if isinstance(payload, dict):
        for key in ("data", "records", "items", "result", "rows"):
            value = payload.get(key)
            if isinstance(value, list):
                return value

        for value in payload.values():
            found = find_records(value)
            if isinstance(found, list):
                return found

    return payload


def make_dataframe(payload: Any) -> pd.DataFrame:
    records = find_records(payload)

    if isinstance(records, list):
        if not records:
            return pd.DataFrame([{"message": "API вернул пустой список"}])
        return pd.json_normalize(records)

    if isinstance(records, dict):
        return pd.json_normalize(records)

    return pd.DataFrame({"value": [str(records)]})


def find_date_column(df: pd.DataFrame) -> str | None:
    for col in ("reportDate", "reporting_date", "date", "period", "Дата"):
        if col in df.columns:
            return col
    return None


def filter_by_year(df: pd.DataFrame, year: int) -> pd.DataFrame:
    date_col = find_date_column(df)

    if not date_col:
        return df

    result = df.copy()
    dates = pd.to_datetime(result[date_col], errors="coerce")
    mask = dates.dt.year == year

    if mask.any():
        return result[mask].copy()

    return df


def make_chart_dataframe(df: pd.DataFrame) -> pd.DataFrame:
    if df is None or df.empty:
        return pd.DataFrame()

    date_col = find_date_column(df)

    if not date_col:
        return pd.DataFrame()

    value_col = None

    if "amount" in df.columns:
        value_col = "amount"
    elif "Значение" in df.columns:
        value_col = "Значение"
    elif "Цена" in df.columns:
        value_col = "Цена"
    else:
        for col in df.columns:
            if col == date_col:
                continue

            numeric = pd.to_numeric(df[col], errors="coerce")
            if numeric.notna().sum() > 0:
                value_col = col
                break

    if not value_col:
        return pd.DataFrame()

    chart_df = df[[date_col, value_col]].copy()
    chart_df[date_col] = pd.to_datetime(chart_df[date_col], errors="coerce")
    chart_df[value_col] = pd.to_numeric(chart_df[value_col], errors="coerce")

    chart_df = chart_df.dropna(subset=[date_col, value_col])

    if chart_df.empty:
        return pd.DataFrame()

    chart_df = (
        chart_df
        .groupby(date_col, as_index=False)[value_col]
        .mean()
        .sort_values(date_col)
    )

    chart_df = chart_df.rename(columns={
        date_col: "Дата",
        value_col: "Значение",
    })

    return chart_df


def add_excel_chart(writer: pd.ExcelWriter, chart_df: pd.DataFrame, title: str) -> None:
    wb = writer.book

    if chart_df.empty or len(chart_df) < 2:
        ws_chart = wb.create_sheet("График")
        ws_chart["A1"] = "Недостаточно данных для построения графика"
        return

    chart_df.to_excel(writer, sheet_name="График_данные", index=False)

    ws_chart = wb.create_sheet("График")

    plt.figure(figsize=(11, 6))
    plt.rcParams["font.family"] = "DejaVu Sans"

    x_labels = chart_df["Дата"].dt.strftime("%Y-%m-%d").tolist()
    x_positions = range(len(x_labels))
    y_values = chart_df["Значение"].tolist()

    plt.plot(
        x_positions,
        y_values,
        color="#0a1a5b",
        linewidth=3,
        marker="o",
        markersize=6,
        label="Значение",
    )

    max_y = max(y_values) if y_values else 0

    for i, (x, y) in enumerate(zip(x_positions, y_values)):
        if i % 2 == 0:
            offset = max_y * 0.035 if max_y else 1
            plt.text(
                x,
                y + offset,
                f"{y:,.1f}",
                ha="center",
                va="bottom",
                fontsize=9,
                color="#000000",
                fontweight="bold",
            )

    plt.title(title, fontsize=14, weight="bold", pad=20)
    plt.grid(True, color="#d9d9d9", linestyle="-", linewidth=0.7, alpha=0.6)
    plt.xlabel("")
    plt.ylabel("")
    plt.xticks(x_positions, x_labels, rotation=30, fontsize=9)
    plt.yticks(fontsize=9)

    if min(y_values) >= 0:
        plt.ylim(bottom=0)

    plt.legend(loc="upper left", fontsize=9, frameon=False)

    plt.text(
        -0.3,
        min(y_values) if y_values else 0,
        "По данным НБ РК / Excel-источников",
        fontsize=9,
        color="#555555",
    )

    plt.text(
        max(len(x_labels) - 3, 0),
        min(y_values) if y_values else 0,
        "TENGENOMIKA",
        fontsize=10,
        color="#0a1a5b",
        weight="bold",
    )

    plt.tight_layout()

    img_data = BytesIO()
    plt.savefig(img_data, format="png", dpi=300, bbox_inches="tight")
    plt.close()
    img_data.seek(0)

    img = Image(img_data)
    ws_chart.add_image(img, "B2")


def style_workbook(writer: pd.ExcelWriter) -> None:
    thin = Side(border_style="thin", color="D9D9D9")
    header_fill = PatternFill("solid", fgColor="EAF0F8")
    header_font = Font(bold=True, color="0A1A5B")

    for sheet_name in writer.sheets:
        ws = writer.sheets[sheet_name]

        for row in ws.iter_rows():
            for cell in row:
                cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
                cell.alignment = Alignment(horizontal="center", vertical="center")

        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill

        for col in ws.columns:
            max_len = 0
            letter = col[0].column_letter

            for cell in col:
                value = "" if cell.value is None else str(cell.value)
                max_len = max(max_len, min(len(value), 60))

            ws.column_dimensions[letter].width = max(12, max_len + 2)

        ws.freeze_panes = "A2"


def generate_report(indicator: dict[str, Any], year: int, payload: Any) -> Path:
    df = make_dataframe(payload)

    if df is None:
        df = pd.DataFrame([{
            "message": "API не вернул данные по выбранному показателю и году"
        }])

    df_before_filter = len(df)

    df = filter_dataframe(df, indicator)

    if df is None:
        df = pd.DataFrame([{
            "message": "После фильтрации данных не осталось"
        }])

    df = filter_by_year(df, year)

    temp_dir = Path(tempfile.gettempdir())
    filename = f"{safe_filename(indicator['name'])}_{year}.xlsx"
    path = temp_dir / filename

    meta = pd.DataFrame([
        {"Поле": "Показатель", "Значение": indicator.get("name", "")},
        {"Поле": "Раздел", "Значение": indicator.get("section", "")},
        {"Поле": "Год", "Значение": year},
        {"Поле": "form_id", "Значение": indicator.get("form_id", "")},
        {"Поле": "API", "Значение": indicator.get("api_url", "")},
        {"Поле": "Источник", "Значение": indicator.get("source_url", "")},
        {"Поле": "Строк до фильтрации", "Значение": df_before_filter},
        {"Поле": "Строк после фильтрации и года", "Значение": len(df)},
    ])

    chart_df = make_chart_dataframe(df)

    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        meta.to_excel(writer, sheet_name="Описание", index=False)
        df.to_excel(writer, sheet_name="Данные", index=False)

        add_excel_chart(
            writer=writer,
            chart_df=chart_df,
            title=str(indicator.get("name", "График")),
        )

        style_workbook(writer)

    return path